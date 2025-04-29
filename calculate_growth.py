import pandas as pd
from pathlib import Path
from datetime import datetime, date, timedelta
import subprocess
import re
from io import StringIO
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

def get_last_7days():
     now = date.today()
     a = []
     for i in range(7):
          day = now - timedelta(days=i)
          a.append(day.strftime("%d-%m-%Y"))
     data = pd.DataFrame(a)

     return data[0].iloc[6],data[0].iloc[5],data[0].iloc[4],data[0].iloc[3],data[0].iloc[2],data[0].iloc[1],data[0].iloc[0]

def get_max_avg_growth(filepath,sheetname):
    date1,date2,date3,date4,date5,date6,date7 = get_last_7days()
    growth_columns = ['growth1','growth2','growth3','growth4','growth5','growth6']
    data = pd.read_excel(Path(filepath),sheet_name=sheetname,usecols=['database',date1,date2,date3,date4,date5,date6,date7])
    last_size = data[data['database'].isnull()]
    data[growth_columns] = data[[date2,date3,date4,date5,date6,date7]] - data[[date1,date2,date3,date4,date5,date6]].values
    growth_values = data[['database','growth1','growth2','growth3','growth4','growth5','growth6']]
    nanrows = growth_values[growth_values['database'].isnull()]
    avg_growth = nanrows[growth_columns].mean(axis=1)
    max_growth = nanrows[growth_columns].max(axis=1)

    return max_growth.iloc[0].round(2), avg_growth.iloc[0].round(2), last_size[date7].iloc[0]

def get_disk_info(host,drive):
    #-m means show the size in megabyte if you want to show in GB use -BG but you need to clean the data with remove G alphabet in the result
    command = f"ssh {host} df -m"
    try:
        disk = subprocess.getoutput(command)
        clean_data = re.sub(r" +",",",disk)
        data = pd.read_csv(StringIO(clean_data),usecols=['Filesystem','1M-blocks','Used','Available','Use%','Mounted'])
        a = data.loc[data['Filesystem'] == drive]
        size = a['1M-blocks'].iloc[0]
        used = a['Used'].iloc[0]
        avail = a['Available'].iloc[0]
        used_pct = a['Use%'].iloc[0]
        #print('size : ', size,'\nused : ', used, '\navail : ', avail, '\nused_pct : ', used_pct )
        
        return a['1M-blocks'].iloc[0], a['Used'].iloc[0], a['Available'].iloc[0], a['Use%'].iloc[0]
        #return data

    except Exception as error:
        print(error)

def how_many_days(filepath,sheetname,insname,instype,detailsheet,avgsize,freespace,lastsize):
    count_days = freespace/avgsize
    #print(count_days)
    wb = load_workbook(filepath,read_only=True)
    if sheetname == wb.sheetnames:
        data = pd.read_excel(filepath,sheetname)
        data.loc[(data['instance_name'] == insname) & (data['instance_type'] == instype), 'days'] = count_days
        data.loc[(data['instance_name'] == insname) & (data['instance_type'] == instype), 'database_size'] 

        with pd.ExcelWriter(filepath,engine='openpyxl',mode='a',if_sheet_exists='replace') as writer:
            data.to_excel(writer,sheet_name=sheetname,index=False)

    else:
        init_data = {"instance_name": [insname],
                      "instance_type": [instype],
                      "days": [round(abs(count_days))],
                      "database_size": [lastsize],
        }
        data = pd.DataFrame(init_data)

        with pd.ExcelWriter(filepath,engine='openpyxl',mode='a',if_sheet_exists='replace') as writer:
            data.to_excel(writer,sheet_name=sheetname,index=False)

     