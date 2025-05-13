import pandas as pd
from db_conn import sqlalchemy_conn,get_variables
from dotenv import load_dotenv
from pathlib import Path
import os
from datetime import datetime, date, timedelta
import argparse
import subprocess
import re
from io import StringIO
from calculate_growth import  get_max_avg_growth, get_disk_info, how_many_days,check_sheet
from openpyxl import load_workbook


def read_excel_size(host,db,xlfile):
    try:
        print("==== read existing xlsx file =====")
        dbsize = pd.read_excel(xlfile,sheet_name=f'{host}_{db}')
        df_columns = len(dbsize.columns)
        #database_size = dbsize.loc[()]
    except Exception as error:
        print(error)
    return dbsize,df_columns

def get_size(host,dbtype,evar):
    try:
        print("==== get new data ====")
        conn = sqlalchemy_conn(host,dbtype,evar)
        now = datetime.today().strftime("%d-%m-%Y")
        if dbtype == 'mysql':
            dbsize = """SELECT table_schema AS "database", 
                        ROUND(SUM(data_length + index_length) / 1024 / 1024) AS "size_mb" 
                        FROM information_schema.TABLES 
                        GROUP BY table_schema;"""
        elif dbtype == 'postgres':
            dbsize = """SELECT
                        pg_database.datname as database,
                        pg_database_size(pg_database.datname)/1024/1024 AS size_mb
                        FROM pg_database;"""
        #print(dbsize)
        df = pd.read_sql(dbsize,con=conn)
        df.rename(columns={"size_mb":f"{now}"},inplace=True)
        df.at['Total',now] = df[now].sum()
    except Exception as error:
        print(error)
    return df

def merge_data(host,dbtype,xlfile,evar):
    
    if os.path.exists(xlfile):
        sheet_name = f'{host}_{dbtype}'
        wb = load_workbook(xlfile,read_only=True)
        #print('merge_data: ',wb.sheetnames)
        if sheet_name in wb.sheetnames:
            try:
                print('file and sheet exist')
                newdata = get_size(host,dbtype,evar)
                xldata = read_excel_size(host,dbtype,xlfile)
                data = pd.merge(xldata[0],newdata,on=['database','database'],how='outer')
                with pd.ExcelWriter(xlfile, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    data.to_excel(writer,sheet_name=f'{host}_{dbtype}',index=False)
                    
                #     return data
            except Exception as error:
                    print(error)
        else:
            try:
                newdata = get_size(host,dbtype,evar)
                with pd.ExcelWriter(xlfile, engine='openpyxl', mode='a') as writer:
                    newdata.to_excel(writer,sheet_name=f'{host}_{dbtype}',index=False)
                #return newdata
            except Exception as error:
                print(error)
    else:
        try:
            print('file not exists')
            newdata = get_size(host,dbtype,evar)
            with pd.ExcelWriter(xlfile, engine='openpyxl', mode='w') as writer:
                newdata.to_excel(writer,sheet_name=f'{host}_{dbtype}',index=False)
            #return newdata
        except Exception as error:
            print(error)
    
def save_detail_space(filepath,detail_sheet,host,dbtype,dsize,dfree,dused,mgrowth,agrowth):
    data = pd.read_excel(filepath,sheet_name=detail_sheet)
    if data.empty:
        print("sheet is empty")
        size_data = {"instance_host": [host],
                    "instance_type": [dbtype],
                    #"disk_drive": ,
                    "disk_size": [dsize],
                    "disk_free_space": [dfree],
                    "disk_used_space": [dused],
                    "max_growth": [mgrowth],
                    "avg_growth": [agrowth]}
        data = pd.DataFrame(size_data)
        return data
    else:
        filter_col = data.loc[(data['instance_host'] == host) & (data['instance_type'] == dbtype)]
        if not filter_col.empty:
            print('data ada')
            filter_col.loc[(filter_col['instance_host'] == host) & (filter_col['instance_type'] == dbtype),'disk_size'] = dsize
            filter_col.loc[(filter_col['instance_host'] == host) & (filter_col['instance_type'] == dbtype),'disk_free_space'] = dfree
            filter_col.loc[(filter_col['instance_host'] == host) & (filter_col['instance_type'] == dbtype),'disk_used_space'] = dused
            filter_col.loc[(filter_col['instance_host'] == host) & (filter_col['instance_type'] == dbtype),'max_growth'] = int(mgrowth)
            filter_col.loc[(filter_col['instance_host'] == host) & (filter_col['instance_type'] == dbtype),'avg_growth'] = int(agrowth)
            return filter_col
        else:
            print('data ga ada')
            size_data = {"instance_host": [host],
                        "instance_type": [dbtype],
                        "disk_size": [dsize],
                        "disk_free_space": [dfree],
                        "disk_used_space": [dused],
                        "max_growth": [mgrowth],
                        "avg_growth": [agrowth]}
            data = pd.DataFrame(size_data)
            return data
    
def get_detail_space(filepath,sheetname,insname,instype):
    try:
        data = pd.read_excel(filepath,sheet_name=sheetname)
        free_space = data.loc[(data['instance_host'] == insname) & (data['instance_type'] == instype), 'disk_free_space'].values[0]
        avg_growth = data.loc[(data['instance_host'] == insname) & (data['instance_type'] == instype), 'avg_growth'].values[0]
        max_growth = data.loc[(data['instance_host'] == insname) & (data['instance_type'] == instype), 'max_growth'].values[0]
        return free_space,avg_growth,max_growth
    except Exception as error:
        print(error)

def sizing_flow(xlfile,envfile):
    load_dotenv(Path(envfile))
    detail_sheet = 'detail_size'
    summary_sheet = 'summary_sheet'
    host = os.environ['HOST'].split(",")
    database = os.environ['DATABASE'].split(",")
    fdata = pd.DataFrame(list(zip(host,database)),columns=['host','database'])
    try:
        check_sheet(detail_sheet,xlfile)
        new_data = []
        for index,row in fdata.iterrows():
            uname, passwd, dbname, hostname, dbdriver, ddrive = get_variables(row['host'],row['database'],envfile)
            sheetname = f'{hostname}_{dbname}'
            
            merge_data(hostname,dbname,xlfile,envfile)

            pd_columns = read_excel_size(hostname,dbname,xlfile)
            if pd_columns[1] > 7:
                max_growth, avg_growth, last_size = get_max_avg_growth(xlfile,sheetname)
                size, used, avail, used_pct = get_disk_info(hostname,ddrive)
                detail_size = save_detail_space(xlfile,detail_sheet,hostname,dbname,size,avail,used,max_growth,avg_growth)
                new_data.append(detail_size)

            else:
                print("== data size is less than 7 days no need to calculated yet ==")
        detail_data = pd.concat(new_data,ignore_index=True)
        with pd.ExcelWriter(xlfile, engine='openpyxl', mode='a',if_sheet_exists='replace') as writer:
            detail_data.to_excel(writer,sheet_name=detail_sheet,index=False)
    
    except Exception as error:
        print(error)


    try:
        check_sheet(summary_sheet,xlfile)
        days_data = []
        for index,row in fdata.iterrows():
            
            uname, passwd, dbname, hostname, dbdriver, ddrive = get_variables(row['host'],row['database'],envfile)
            sheetname = f'{hostname}_{dbname}'
            max_growth, avg_growth, last_size = get_max_avg_growth(xlfile,sheetname)
            free_space,avg_growth,max_growth = get_detail_space(xlfile,detail_sheet,hostname,dbname)
            get_days = how_many_days(xlfile,summary_sheet,hostname,dbname,detail_sheet,avg_growth,free_space,last_size)
            days_data.append(get_days)
        summary_data = pd.concat(days_data,ignore_index=True)
        print(summary_data)
        with pd.ExcelWriter(xlfile, engine='openpyxl', mode='a',if_sheet_exists='replace') as writer:
            summary_data.to_excel(writer,sheet_name=summary_sheet,index=False)

    except Exception as error:
        print(error)

if __name__ == '__main__':
    
    parser = argparse.ArgumentParser(prog="db sizing",description="db sizing")
    parser.add_argument('-f','--excel-file',required=True, help="excel file that used to save the database size history")
    parser.add_argument('-e','--env-file',required=True,help="env file for configuration")
    args = vars(parser.parse_args())
    sizing_flow(args['excel_file'],args['env_file'])
    
    